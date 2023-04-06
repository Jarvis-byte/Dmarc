from email import encoders
from email.mime.base import MIMEBase

from googleapiclient.http import MediaFileUpload
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
import csv
from datetime import datetime
import requests
import json
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from datetime import datetime
from PolicyChangeEncoder import PolicyChangeEncoder
from google.oauth2.service_account import Credentials
from googleapiclient.errors import HttpError
from googleapiclient.discovery import build

# to = ['df-rtps-emailsecurityteam@randstad.com']

to = ['arka.mazumder@randstad.com', 'saravanakumar.g@randstad.com']
cc = ['balaji.natarajan@randstad.com']
# emailFrom = 'dmarc-no-reply@randstad.com'
report_date = datetime.now().strftime('%A, %d. %B %Y %R')
subject = f"Dmarc Data Ingestion - OpCo, SPF & DMARC Lookup: {report_date} (testing)"
body = None
SPREADSHEET_ID = '188lKNuzMXplzkO6UpZLsFOHx-NgDXA3lfZpvkSV-neE'

# Replace with the path to your service account JSON key file
SERVICE_ACCOUNT_FILE = 'gis-bcdr-ops-fe35-84aa16f1f12c.json'

# Authenticate and build the service object
creds = Credentials.from_service_account_file(
    SERVICE_ACCOUNT_FILE,
    scopes=['https://www.googleapis.com/auth/drive', 'https://www.googleapis.com/auth/spreadsheets']
)

service = build('drive', 'v3', credentials=creds)


class ExportToExcel:
    def export(self, list, Dmarc_Policy_Change):
        file_path = "output.xlsx"
        workbook = openpyxl.load_workbook(filename=file_path)
        sheet = workbook['Sheet']
        rows = sheet.max_row
        # print(f'{rows} is the DMARC Policy')
        for r in range(0, rows - 1):
            row = sheet[r + 2]

            for c in range(5, 9):
                cell = row[c]
                if c == 6:
                    cell.value = list[r].spf_policy
                elif c == 7:
                    cell.value = list[r].dmarc_policy
                elif c == 8:
                    cell.value = list[r].spf_count

        try:
            workbook.save(filename=file_path)
            file_path = file_path
            workbook.save(filename=file_path)
            print("output.xlsx has been created successfully")
            file_metadata = {'name': 'output'}
            media = MediaFileUpload('output.xlsx',
                                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                    resumable=True)

            updated_file = service.files().update(fileId=SPREADSHEET_ID, body=file_metadata, media_body=media,
                                                  fields='id').execute()
            print('Modified Google Sheet uploaded to Google Drive')
            wb = openpyxl.load_workbook('output.xlsx')

            # Select the active worksheet
            ws = wb.active

            # Get the header row
            header = [cell.value for cell in ws[1]]

            # Get the data rows
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                data.append(dict(zip(header, row)))

            # Convert the data to JSON format
            json_data = json.dumps(data, indent=4)

            # Write the JSON data to a file
            with open('Domain_info.json', 'w') as f:
                f.write(json_data)

            print("JSON data saved to Domain_info.json")
            now = datetime.now()

            current_time = now.strftime("%H:%M:%S")
            print("End Time =", current_time)

            # Adding to splunk.
            # uri = "https://http-inputs-randstad.splunkcloud.com/services/collector"
            # # Enter the original token ->35B8DBEA-2E3C-41DF-A318-EE46154B7AEA
            # splunk_token = "35B8DBEA-2E3C-41DF-A318-EE46154B7AEA"
            # out_path = "Domain_info.json"
            # host_name = "dmarc.orgV2"
            # src = "dmarc:lookupV2"
            #
            # with open(out_path, "r") as f:
            #     input_objects = json.load(f)
            #
            # for input_object in input_objects:
            #     body = json.dumps({"event": input_object, "host": host_name, "source": src})
            #     # print("Body:", body)
            #     headers = {"Authorization": f"Splunk {splunk_token}"}
            #     response = requests.post(uri, data=body, headers=headers)
            #
            #     if response.text != "Success":
            #         print(response.text)

            # Code foe finding the previous day data with present day data. It will only be possible if we store the output excel sheet, otherwise not possible
            # Finding what all change have happened previous day+

            # Checking for change in policy
            Total_downgrade = 0
            Change_spf = 0
            transformation_change = 0
            Change_None_to_Q = 0
            Change_None_to_R = 0
            Change_Q_to_R = 0
            Change_R_to_Q = 0
            Change_R_to_None = 0
            Change_Q_to_None = 0

            for a in range(0, len(Dmarc_Policy_Change)):
                if (Dmarc_Policy_Change[a].dmarc_policy_old.lower() == "none" and Dmarc_Policy_Change[
                    a].dmarc_policy_new.lower() == "quarantine"):
                    Change_None_to_Q += 1

                if Dmarc_Policy_Change[a].dmarc_policy_old.lower() == "none" and Dmarc_Policy_Change[
                    a].dmarc_policy_new.lower() == "reject":
                    print('Yes')
                    Change_None_to_R += 1

                if Dmarc_Policy_Change[a].dmarc_policy_old.lower() == "quarantine" and Dmarc_Policy_Change[
                    a].dmarc_policy_new.lower() == "reject":
                    Change_Q_to_R += 1

                #     Downgrade count
                if (Dmarc_Policy_Change[a].dmarc_policy_old.lower() == "reject" and Dmarc_Policy_Change[
                    a].dmarc_policy_new.lower() == "quarantine"):
                    Change_R_to_Q += 1
                    Total_downgrade += 1

                if (Dmarc_Policy_Change[a].dmarc_policy_old.lower() == "reject" and Dmarc_Policy_Change[
                    a].dmarc_policy_new.lower() == "none"):
                    Change_R_to_None += 1
                    Total_downgrade += 1
                if (Dmarc_Policy_Change[a].dmarc_policy_old.lower() == "quarantine" and Dmarc_Policy_Change[
                    a].dmarc_policy_new.lower() == "none"):
                    Change_Q_to_None += 1
                    Total_downgrade += 1

                if (Dmarc_Policy_Change[a].spf_policy_change_old == None and Dmarc_Policy_Change[
                    a].spf_policy_change_new != None) or (
                        Dmarc_Policy_Change[a].spf_policy_change_old.lower() != Dmarc_Policy_Change[
                    a].spf_policy_change_new.lower()):
                    Change_spf += 1

                if (Dmarc_Policy_Change[a].transformation_old.lower() != Dmarc_Policy_Change[
                    a].transformation_new.lower()):
                    transformation_change += 1

            print(f'change in None to Q - > {Change_None_to_Q}')
            print(f'Change in None to R -> {Change_None_to_R}')
            print(f'Change in Q to R -> {Change_Q_to_R}')
            print(f'Change in SPF - > {Change_spf}')

            print(f'Change in Downgrade - > {Total_downgrade}')
            print(f'Change in Downgrade -> Reject To Q  - > {Change_R_to_Q}')
            print(f'Change in Downgrade Reject to None - > {Change_R_to_None}')
            print(f'Change in Downgrade Q to N - > {Change_Q_to_None}')
            print(f'Change in Transformation - > {transformation_change}')

            json_string = json.dumps(Dmarc_Policy_Change, cls=PolicyChangeEncoder)
            with open('Change_Details.json', 'w') as f:
                # Write the JSON string to the file
                f.write(json_string)

            # create email message
            # emailFrom = 'arkamazumder0@gmail.com'
            # body = "The attached JSON file that includes the details of OpCo, SPF, and DMARC policy has been successfully ingested into Splunk."
            # msg = MIMEMultipart()
            # msg['To'] = ', '.join(to)
            # msg['Cc'] = ', '.join(cc)
            # msg['Subject'] = subject
            # msg['From'] = ''
            # msg.attach(MIMEText(body))
            # username = 'arkamazumder0@gmail.com'
            # password = 'vgcsinjjmzkgufjz'
            #
            # # with open('file_path2', 'rb') as f:
            # #     attachment = MIMEBase('application', 'octet-stream')
            # #     attachment.set_payload(f.read())
            # #     encoders.encode_base64(attachment)
            # #     attachment.add_header('Content-Disposition', 'attachment', filename=filename)
            # #     msg.attach(attachment)
            # with open('Domain_info.json', 'rb') as f:
            #     attachment = MIMEBase('application', 'octet-stream')
            #     attachment.set_payload(f.read())
            #     encoders.encode_base64(attachment)
            #     attachment.add_header('Content-Disposition', 'attachment', filename='Total Domain Information.json')
            #     msg.attach(attachment)
            #     try:
            #         # server = smtplib.SMTP('smtpeu.randstad.gis', 25)
            #         server = smtplib.SMTP('smtp.gmail.com', 587)
            #         server.ehlo()
            #         server.starttls()
            #         # not req
            #         server.login(username, password)
            #         server.sendmail(emailFrom, to + cc, msg.as_string())
            #         server.close()
            #         print("Email sent successfully")
            #     except Exception as e:
            #         print(f'Error: {str(e)}')

            # Create the table data
            dataUpgrade = [
                [Change_None_to_Q, Change_None_to_R, Change_Q_to_R]]
            dataDowngrade = [[Change_R_to_Q, Change_R_to_None, Change_Q_to_None]]
            # for change in Dmarc_Policy_Change:
            #     data.append(
            #         [change.domain_name, change.dmarc_policy_old, change.dmarc_policy_new, change.spf_policy_change_old,
            #          change.spf_policy_change_new, change.transformation_old, change.transformation_new])

            # Create the HTML table from the data for upgrade
            table_html1 = '<table style="border-collapse: collapse; border: 2px solid black;">'
            # Add header row
            table_html1 += '<tr><th style="border: 2px solid black; padding: 8px;">None to Quarantine</th><th style="border: 2px solid black; padding: 8px;">None to Reject</th><th style="border: 2px solid black; padding: 8px;">Quarantine to Reject</th>'
            # Add data rows
            for row in dataUpgrade:
                table_html1 += '<tr>'
                for value in row:
                    table_html1 += f'<td style="border: 2px solid black; padding: 8px;">{str(value)}</td>'
                table_html1 += '</tr>'
            table_html1 += '</table>'

            # Create the HTML table from the data for downgrade
            table_html2 = '<table style="border-collapse: collapse; border: 2px solid black;">'
            # Add header row
            table_html2 += '<tr><th style="border: 2px solid black; padding: 8px;">Reject to Quarantine</th><th style="border: 2px solid black; padding: 8px;">Reject to None</th><th style="border: 2px solid black; padding: 8px;">Quarantine to None</th>'
            # Add data rows
            for row in dataDowngrade:
                table_html2 += '<tr>'
                for value in row:
                    table_html2 += f'<td style="border: 2px solid black; padding: 8px;">{str(value)}</td>'
                table_html2 += '</tr>'
            table_html2 += '</table>'

            to = ['arka.mazumder@randstad.com', 'saravanakumar.g@randstad.com', 'balaji.natarajan@randstad.com']
            # cc = ['balaji.natarajan@randstad.com']

            # 'saravanakumar.g@randstad.com','balaji.natarajan@randstad.com'
            emailFrom = 'arkamazumder0@gmail.com'
            reportDate = datetime.now().strftime("%A, %d. %B %Y %R")
            if Total_downgrade == 0:
                emailSubject = f"Dmarc Data Ingestion - OpCo, SPF & DMARC Lookup: {reportDate} (testing with DUMMY data)"
            else:
                emailSubject = f"ALERT !!! POLICY DOWNGRADED. Dmarc Data Ingestion - OpCo, SPF & DMARC Lookup: {reportDate}. (testing with DUMMY data)"
            emailMessage = MIMEMultipart()
            emailMessage['From'] = emailFrom
            emailMessage['To'] = ', '.join(to)
            emailMessage['Subject'] = emailSubject
            if len(Dmarc_Policy_Change) == 0 and Total_downgrade == 0:
                emailMessage.attach(MIMEText(
                    f'<html><body>The attached JSON file that includes the details of OpCo, SPF, and DMARC policy has been successfully ingested into Splunk.<br><br> No Policy Has been Changed Yesterday.',
                    'html'

                ))
            else:
                emailMessage.attach(MIMEText(
                    f'<html><body>The attached JSON file that includes the details of OpCo, SPF, and DMARC policy has been successfully ingested into Splunk.<br><br>Any Policy change for the domain/domains are below:-<br><br>{table_html1}<br><br><strong>Alert</strong> - Total Policy Downgrade:\t '
                    f'{Total_downgrade}<br><br>{table_html2}<br><br>For more information please find attach</body></html>',
                    'html'

                ))

            out = "Domain_info.json"
            attachment = MIMEApplication(open(out, "rb").read(), _subtype="json")
            attachment.add_header('Content-Disposition', 'attachment', filename=os.path.basename(out))
            emailMessage.attach(attachment)
            out2 = "Change_Details.json"
            attachment = MIMEApplication(open(out2, "rb").read(), _subtype="json")
            attachment.add_header('Content-Disposition', 'attachment', filename=os.path.basename(out2))
            emailMessage.attach(attachment)
            # try:
            #     smtpClient = smtplib.SMTP_SSL(smtpServer, smtpServerPort)
            #     smtpClient.ehlo()
            #
            #     smtpClient.sendmail(emailFrom, emailTo, emailMessage.as_string())
            #     print(emailFrom)
            #     smtpClient.login(emailFrom, 'vgcsinjjmzkgufjz')
            #     # vgcsinjjmzkgufjz
            #     smtpClient.close()
            # except Exception as e:
            #     print(f'Error Mail Sending:- {str(e)}')
            with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
                smtp.ehlo()
                smtp.login(emailFrom, 'vgcsinjjmzkgufjz')
                smtp.sendmail(emailFrom, to, emailMessage.as_string())
                smtp.close()


        except Exception as e:
            print(e)
        finally:
            workbook.close()
