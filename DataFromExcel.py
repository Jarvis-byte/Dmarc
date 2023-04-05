import multiprocessing
import concurrent.futures
import openpyxl
from ExportToExcel import ExportToExcel
from PolicyChange import PolicyChange
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import multiprocessing as mp
from Data import Data
from HttpHandler import HttpHandler


class DataFromExcel:
    def readData(self, excelPath):

        workbook = openpyxl.load_workbook(excelPath)
        sheet = workbook['Sheet']
        data_list = []
        Old_data_List = []

        rows = sheet.max_row
        print("Rows\t" + str(rows))
        ws = workbook['Sheet']

        # Iterate through rows and columns
        for row in range(2, ws.max_row + 1):
            data = Data()
            for col in range(1, ws.max_column + 1):

                cell_value = ws.cell(row=row, column=col).value
                if col == 1:
                    data.opco_code = cell_value
                if col == 2:
                    data.domain_name = cell_value
                if col == 3:
                    data.transformation = cell_value
                if col == 7:
                    data.spf_policy = cell_value
                if col == 8:
                    data.dmarc_policy = cell_value

            data_list.append(data)
            Old_data_List = [data.copy() for data in data_list]

        # print(f'Dmarc for old -> {Old_data_List[2].dmarc_policy}')
        part_size = len(data_list) // 8
        start_index = 0
        part_indices = []
        for i in range(8):
            end_index = start_index + part_size
            if i == 7:
                end_index = len(data_list)  # for the last part
            part_indices.append((start_index, end_index))
            start_index = end_index


        with concurrent.futures.ThreadPoolExecutor() as executor:
            f1 = executor.submit(HttpHandler().get_request, data_list, part_indices[0][0], part_indices[0][1])
            f2 = executor.submit(HttpHandler().get_request, data_list, part_indices[1][0], part_indices[1][1])
            f3 = executor.submit(HttpHandler().get_request, data_list, part_indices[2][0], part_indices[2][1])
            f4 = executor.submit(HttpHandler().get_request, data_list, part_indices[3][0], part_indices[3][1])
            f5 = executor.submit(HttpHandler().get_request, data_list, part_indices[4][0], part_indices[4][1])
            f6 = executor.submit(HttpHandler().get_request, data_list, part_indices[5][0], part_indices[5][1])
            f7 = executor.submit(HttpHandler().get_request, data_list, part_indices[6][0], part_indices[6][1])
            f8 = executor.submit(HttpHandler().get_request, data_list, part_indices[7][0], part_indices[7][1])
            # + f2.result() + f3.result() + f4.result() + f5.result() + f6.result() + f7.result() + f8.result()
            results = f1.result() + f2.result() + f3.result() + f4.result() + f5.result() + f6.result() + f7.result() + f8.result()

        # Find the policy change

            Dmarc_Policy_Change = []

            for a in range(0, len(Old_data_List)):
                check = False
                policyclass = PolicyChange()
                policyclass.domain_name = Old_data_List[a].domain_name
                policyclass.opco_code = Old_data_List[a].opco_code

                old_dmarc_policy = Old_data_List[a].dmarc_policy
            # print(old_dmarc_policy)
                final_old_p = None

                if old_dmarc_policy != "No DMARC Record" and old_dmarc_policy != None:
                    split1 = old_dmarc_policy.split(";")
                    split2 = split1[1].split("=")
                    final_old_p = split2[1]
                # print(f'Final Old P -> {final_old_p}')

                new_dmarc_policy = results[a].dmarc_policy
                final_new_p = None
                if new_dmarc_policy != "No DMARC Record":
                    split1 = new_dmarc_policy.split(";")
                    split2 = split1[1].split("=")
                    final_new_p = split2[1]
                # print(f'Final New P -> {final_new_p}')
                if old_dmarc_policy != None and (final_old_p != final_new_p):
                    policyclass.dmarc_policy_old = final_old_p
                    policyclass.dmarc_policy_new = final_new_p
                    check = True
                if Old_data_List[a].spf_policy != None and Old_data_List[a].spf_policy != results[a].spf_policy:
                    policyclass.spf_policy_change_old = Old_data_List[a].spf_policy
                    policyclass.spf_policy_change_new = results[a].spf_policy
                    check = True
                if Old_data_List[a].transformation != None and Old_data_List[a].transformation != results[
                    a].transformation:
                    policyclass.transformation_old = Old_data_List[a].transformation
                    policyclass.transformation_new = results[a].transformation
                    check = True
                if check:
                    Dmarc_Policy_Change.append(policyclass)

            # print(f'The change for quarantine is -> {cnt_Quarantine} and for Reject is -> {cnt_Reject}')
            export_to_excel = ExportToExcel()
            export_to_excel.export(results, Dmarc_Policy_Change)
        # for a in range(0, len(results)):
        #     print(f'Policy in result -> {results[a].domain_name} -> \t {results[a].dmarc_policy}')

        # result_queue1 = mp.Queue()
        # result_queue2 = mp.Queue()
        # result_queue3 = mp.Queue()
        # result_queue4 = mp.Queue()
        # result_queue5 = mp.Queue()
        #
        # p1 = mp.Process(target=HttpHandler().get_request, args=(data_list, 0, 3, result_queue1,))
        # p2 = mp.Process(target=HttpHandler().get_request, args=(data_list, 3, 8, result_queue2,))
        # p3 = mp.Process(target=HttpHandler().get_request, args=(data_list, 8, len(data_list), result_queue3,))
        # # p4 = mp.Process(target=HttpHandler().get_request, args=(data_list, 1200, 1600, result_queue4,))
        # # p5 = mp.Process(target=HttpHandler().get_request, args=(data_list, 1600, len(data_list), result_queue5,))
        #
        #
        # p1.start()
        # p2.start()
        # p3.start()
        # #p4.start()
        # #p5.start()
        #
        #
        # p1.join()
        # p2.join()
        # p3.join()
        # #p4.join()
        # #p5.join()
        # print(p3.is_alive())

        # returned_data_list = []
        # while not result_queue1.empty():
        #     returned_data_list += result_queue1.get()
        # print(p1.is_alive())
        # result_queue1.close()
        # result_queue1.join_thread()
        #
        # while not result_queue2.empty():
        #     returned_data_list += result_queue2.get()
        # result_queue2.close()
        # result_queue2.join_thread()
        #
        # while not result_queue3.empty():
        #     returned_data_list += result_queue3.get()
        # result_queue3.close()
        # result_queue3.join_thread()
        #
        # while not result_queue4.empty():
        #     returned_data_list += result_queue4.get()
        # result_queue4.close()
        # result_queue4.join_thread()
        #
        # while not result_queue5.empty():
        #     returned_data_list += result_queue5.get()
        # result_queue5.close()
        # result_queue5.join_thread()
        #
        #
        #
        # export_to_excel = ExportToExcel()
        # export_to_excel.export(returned_data_list)
        # http_handler = HttpHandler()
        # http_handler.get_request(data_list[:500])
